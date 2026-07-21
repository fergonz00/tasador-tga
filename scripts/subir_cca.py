# -*- coding: utf-8 -*-
"""
Sube el .xlsx del CCA a la pestaña que lee el tasador, sin pasar por el
File → Import de Google Sheets.

Va contra el Apps Script del simulador VWFS (action `cargarCCA`), que es el
que tiene autorización de escritura sobre esa planilla. El script hace backup
de la pestaña anterior antes de pisarla y conserva el gid.

Credenciales en C:\\proyectos\\.secrets\\simulador-vwfs.env (fuera de git).

Uso:  python subir_cca.py cca_julio_2026.xlsx cca_precios_julio_2026
"""
import sys, json, os, urllib.request

ENV = r"C:\proyectos\.secrets\simulador-vwfs.env"


def env():
    out = {}
    for line in open(ENV, encoding="utf-8"):
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            out[k.strip()] = v.strip()
    return out


def leer_xlsx(path):
    from openpyxl import load_workbook
    ws = load_workbook(path, data_only=True).active
    filas = []
    for row in ws.iter_rows(values_only=True):
        filas.append(["" if c is None else c for c in row])
    return filas


def main():
    xlsx = sys.argv[1]
    nombre = sys.argv[2] if len(sys.argv) > 2 else None
    cfg = env()
    filas = leer_xlsx(xlsx)
    print(f"{len(filas)} filas x {len(filas[0])} columnas desde {os.path.basename(xlsx)}")

    payload = {"token": cfg["PP_TOKEN"], "action": "cargarCCA", "valores": filas}
    if nombre:
        payload["nombre"] = nombre
    data = json.dumps(payload).encode("utf-8")
    print(f"POST {len(data)/1024:.0f} KB -> {cfg['CCA_EXEC_URL'][:60]}...")

    req = urllib.request.Request(cfg["CCA_EXEC_URL"], data=data,
                                 headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=600) as r:
        resp = json.loads(r.read().decode("utf-8"))
    print(json.dumps(resp, ensure_ascii=False, indent=2))
    if not resp.get("ok"):
        sys.exit(1)


if __name__ == "__main__":
    main()
